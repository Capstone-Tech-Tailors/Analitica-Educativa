"""Genera las planillas de anotación (una por anotador) a partir de la muestra estratificada.

Salida: anotacion/anotacion_A1.xlsx, _A2.xlsx, _A3.xlsx

Diseño (ver PROTOCOLO_ANOTACION.md):
  - Bloque EVAL  : los 3 anotadores etiquetan los MISMOS ítems -> mide acuerdo y forma el gold.
  - Bloque TRAIN : se reparte en tercios disjuntos -> triplica los datos de entrenamiento.
  - Anclas       : ítems de etiqueta obvia, para detectar anotadores descuidados.
  - Repetidos    : ítems duplicados de forma encubierta, para medir autoconsistencia.
  - Orden aleatorizado independiente por anotador; claves opacas; sin predicción del modelo.

Uso:  python anotacion/generar_planillas.py
"""
from __future__ import annotations

import hashlib
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = next(p for p in [Path.cwd(), *Path.cwd().parents] if (p / "src").is_dir())
sys.path.insert(0, str(RAIZ))
from src import rutas  # noqa: E402

SEMILLA_BASE = 20260728
ANOTADORES = ["A1", "A2", "A3"]
ETIQUETAS = ["POS", "NEU", "NEG"]
MOTIVOS = ["elogio_tibio", "critica_suave", "mixto_adversativa",
           "descriptivo", "muy_corto", "sin_sentido", "otro"]

# Ítems ancla: etiqueta trivial bajo la rúbrica, sin adversativas ni atenuadores.
ANCLAS = [
    ("Excelente profesor, muy claro.", "POS"),
    ("Explica muy bien los temas.", "POS"),
    ("Los talleres son muy útiles para aprender.", "POS"),
    ("La clase es muy dinámica y entretenida.", "POS"),
    ("Califica de forma injusta.", "NEG"),
    ("No explica bien, me siento perdido.", "NEG"),
    ("La clase es muy aburrida y monótona.", "NEG"),
    ("Muy mal profesor.", "NEG"),
]
N_REPETIDOS = 8          # ítems duplicados encubiertos (autoconsistencia)
SEPARACION_MIN = 60      # filas mínimas entre un repetido y su original

GRIS = PatternFill("solid", fgColor="EFEFEA")
CABECERA = PatternFill("solid", fgColor="2A78D6")


def clave(anotador: str, texto: str) -> str:
    """Clave opaca y estable: impide cotejar planillas entre anotadores."""
    h = hashlib.sha1(f"{anotador}|{texto}".encode()).hexdigest()[:6].upper()
    return f"{anotador}-{h}"


def construir_filas(muestra: pd.DataFrame, anotador: str, idx: int) -> pd.DataFrame:
    """Ítems que ve un anotador: EVAL completo + su tercio de TRAIN + anclas + repetidos."""
    ev = muestra[muestra["bloque"] == "EVAL"]["rep"].tolist()
    tr = muestra[muestra["bloque"] == "TRAIN"]["rep"].tolist()
    tercio = tr[idx::len(ANOTADORES)]                       # partición disjunta

    filas = [{"texto": t, "tipo": "eval"} for t in ev]
    filas += [{"texto": t, "tipo": "train"} for t in tercio]
    filas += [{"texto": t, "tipo": "ancla", "esperada": e} for t, e in ANCLAS]

    rng = pd.Series(ev).sample(N_REPETIDOS, random_state=SEMILLA_BASE + idx).tolist()
    filas += [{"texto": t, "tipo": "repetido"} for t in rng]

    df = pd.DataFrame(filas).sample(frac=1, random_state=SEMILLA_BASE + idx).reset_index(drop=True)

    # Separar cada repetido de su original al menos SEPARACION_MIN filas
    for t in rng:
        pos = df.index[df["texto"] == t].tolist()
        if len(pos) == 2 and pos[1] - pos[0] < SEPARACION_MIN:
            fila = df.loc[pos[1]]
            df = df.drop(index=pos[1]).reset_index(drop=True)
            destino = min(pos[0] + SEPARACION_MIN, len(df))
            df = pd.concat([df.iloc[:destino], pd.DataFrame([fila]), df.iloc[destino:]],
                           ignore_index=True)

    df.insert(0, "orden", range(1, len(df) + 1))
    df["anotador"] = anotador
    df["clave"] = [clave(anotador, f"{o}|{t}") for o, t in zip(df["orden"], df["texto"])]
    return df


def escribir_xlsx(df: pd.DataFrame, anotador: str, destino: Path) -> Path:
    wb = Workbook()

    # --- Hoja de instrucciones ---
    ins = wb.active
    ins.title = "INSTRUCCIONES"
    texto = [
        ("Anotación de sentimiento · Analítica Educativa", True),
        ("", False),
        (f"Anotador: {anotador}   ·   Ítems: {len(df)}   ·   Tiempo estimado: ~35 minutos", False),
        ("", False),
        ("1. Lee primero la rúbrica completa (RUBRICA_ANOTACION.md). Son 5 minutos.", False),
        ("2. Ve a la hoja ANOTACION y rellena SOLO las columnas de fondo blanco.", False),
        ("3. Trabaja de arriba a abajo. NO ordenes ni filtres la hoja.", False),
        ("4. Elige la etiqueta del desplegable. No dejes ninguna fila vacía.", False),
        ("5. Si dudas más de 15 segundos: decide, marca dificil = 1 y continúa.", False),
        ("6. NO consultes con los otros anotadores: la independencia es obligatoria.", False),
        ("7. Hazlo de una sola sentada y devuelve el archivo sin renombrarlo.", False),
        ("", False),
        ("Criterio central: ¿qué haría el coordinador académico al leer el comentario?", True),
        ("   Reconocer al docente -> POS", False),
        ("   Nada, no hay información accionable -> NEU", False),
        ("   Mirarlo de cerca / acompañarlo -> NEG", False),
        ("", False),
        ("Recuerda: 'estuvo bien' es cortesía mínima (NEU), y una crítica suave", False),
        ("('se podría mejorar') sigue siendo una crítica (NEG).", False),
    ]
    for i, (linea, negrita) in enumerate(texto, start=1):
        c = ins.cell(row=i, column=1, value=linea)
        c.font = Font(bold=negrita, size=12 if negrita and i == 1 else 11)
    ins.column_dimensions["A"].width = 95

    # --- Hoja de anotación ---
    ws = wb.create_sheet("ANOTACION")
    cols = ["anotador", "orden", "clave", "comentario", "etiqueta", "dificil", "motivo", "nota"]
    ws.append(cols)
    for c in ws[1]:
        c.fill = CABECERA
        c.font = Font(bold=True, color="FFFFFF")

    for _, r in df.iterrows():
        ws.append([r["anotador"], r["orden"], r["clave"], r["texto"], None, 0, None, None])

    anchos = {"A": 10, "B": 8, "C": 14, "D": 62, "E": 11, "F": 9, "G": 20, "H": 30}
    for col, w in anchos.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "D2"

    n = len(df) + 1
    dv_etq = DataValidation(type="list", formula1='"' + ",".join(ETIQUETAS) + '"',
                            allow_blank=False, showErrorMessage=True,
                            errorTitle="Etiqueta no válida", error="Elige POS, NEU o NEG.")
    dv_dif = DataValidation(type="list", formula1='"0,1"', allow_blank=False,
                            showErrorMessage=True)
    dv_mot = DataValidation(type="list", formula1='"' + ",".join(MOTIVOS) + '"', allow_blank=True)
    for dv, letra in ((dv_etq, "E"), (dv_dif, "F"), (dv_mot, "G")):
        ws.add_data_validation(dv)
        dv.add(f"{letra}2:{letra}{n}")

    # Bloquear lo pre-rellenado, dejar editable lo que anota la persona
    for fila in ws.iter_rows(min_row=2, max_row=n, max_col=len(cols)):
        for c in fila:
            editable = c.column_letter in ("E", "F", "G", "H")
            c.protection = Protection(locked=not editable)
            if not editable:
                c.fill = GRIS
            if c.column_letter == "D":
                c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.protection.sheet = True
    ws.protection.password = "techtailors"
    ws.protection.sort = ws.protection.autoFilter = False
    ws.protection.insertRows = ws.protection.deleteRows = False

    ruta = destino / f"anotacion_{anotador}.xlsx"
    wb.save(ruta)
    return ruta


def main() -> None:
    muestra = pd.read_csv(rutas.datos_processed("muestra_anotacion.csv"))
    destino = RAIZ / "anotacion"
    destino.mkdir(exist_ok=True)

    print(f"Muestra: {len(muestra)} ítems "
          f"(EVAL {int((muestra.bloque=='EVAL').sum())} · TRAIN {int((muestra.bloque=='TRAIN').sum())})\n")

    clave_maestra = []
    for i, a in enumerate(ANOTADORES):
        df = construir_filas(muestra, a, i)
        ruta = escribir_xlsx(df, a, destino)
        clave_maestra.append(df.assign(anotador=a))
        print(f"  {ruta.name:22} {len(df):>4} ítems  "
              f"(eval {int((df.tipo=='eval').sum())} · train {int((df.tipo=='train').sum())} · "
              f"ancla {int((df.tipo=='ancla').sum())} · repetido {int((df.tipo=='repetido').sum())})")

    # Clave maestra: SOLO para el coordinador (identifica anclas y repetidos)
    km = pd.concat(clave_maestra, ignore_index=True)
    ruta_km = destino / "_clave_maestra.csv"
    km.to_csv(ruta_km, index=False)
    print(f"\n  {ruta_km.name} -> uso exclusivo del coordinador (no compartir con los anotadores)")
    print("\nSiguiente paso: repartir cada XLSX a su anotador junto con RUBRICA_ANOTACION.md")


if __name__ == "__main__":
    main()
